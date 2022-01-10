#!/usr/bin/env python3
import openpyxl
from openpyxl.utils import cell, get_column_letter, column_index_from_string
import os
import glob
from os import listdir
from os.path import isfile, join
import pprint

print(f"-- testing openpyxl version {openpyxl.__version__}")

# sheets_dir = 'sheets'
# onlyfiles = [f for f in listdir('sheets') if isfile(join(mypath, f))]

script_path = os.path.realpath(__file__)
script_dir = os.path.dirname(script_path)
swc_dir = os.path.join(script_dir, "sheets")

# every file that ends with '.xlsx' under "sheets" dir
# https://stackoverflow.com/questions/18394147/how-to-do-a-recursive-sub-folder-search-and-return-files-in-a-list
files = glob.glob(swc_dir + '/**/*.xlsx', recursive=True)

total = 0
dd = {}

for f in files:
    try:
        print(f)
        wb = openpyxl.load_workbook(f)
        sheet = wb[wb.sheetnames[0]]
        
        # print(sheet.min_row, sheet.max_row, sheet.min_column, sheet.max_column)
        # print(get_column_letter(sheet.min_column), get_column_letter(sheet.max_column)) 
        field_start = get_column_letter(sheet.min_column) + str(sheet.min_row)
        field_end = get_column_letter(sheet.max_column) + str(sheet.max_row)
        # print(field_start, '--', field_end)

        week_nr = None
        week_total = 0

        for rowOfCellObjects in sheet[field_start:field_end]:
            for cellObj in rowOfCellObjects:
                if cellObj.value is not None:
                    if cellObj.value == "S&O - AGV":
                        coord_curr = cellObj.coordinate
                        coord_next = coord_curr.replace(get_column_letter(cellObj.col_idx), get_column_letter(cellObj.col_idx+1))
                        # print(coord_curr, coord_next, sheet[coord_next].value)
                        total += sheet[coord_next].value
                        week_total += sheet[coord_next].value
                    elif str(cellObj.value).startswith("weeknr:"):
                        week_nr = [int(s) for s in str(cellObj.value).split() if s.isdigit()]
                        week_nr = week_nr[0]

        print(f"Week {week_nr}: {week_total}")
        # dd["week"+str(week_nr)] = week_total
        dd[week_nr] = week_total
        
        # for row in range(sheet.max_row):
            # for col in range(sheet.max_column):
                # val = 


        
    except Exception as e:
        print("-----")
        print(f)
        print("Error: ", e)
        print("######")

pprint.pprint(dd)
print(f"Total:{total}")

