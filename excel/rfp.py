#!/usr/bin/env python3
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import xlrd
import glob
import argparse
from os.path import exists, isdir
import pprint

def convert_workbook_xlrd_to_openpyxl(xlrd_workbook):
    # Create a new openpyxl workbook
    openpyxl_workbook = Workbook()

    # Iterate through the sheets in the xlrd workbook
    for sheet_index in range(xlrd_workbook.nsheets):
        xlrd_sheet = xlrd_workbook.sheet_by_index(sheet_index)
        sheet_name = xlrd_sheet.name

        # Create corresponding sheet in the openpyxl workbook
        openpyxl_sheet = openpyxl_workbook.create_sheet(title=sheet_name)

        # Iterate through the cells in the xlrd sheet
        for row_index in range(xlrd_sheet.nrows):
            for col_index in range(xlrd_sheet.ncols):
                cell_value = xlrd_sheet.cell_value(row_index, col_index)
                cell_format = xlrd_sheet.cell_xf_index(row_index, col_index)
                
                # Copy the values
                openpyxl_sheet.cell(row=row_index+1, column=col_index+1).value = cell_value # str(cell_value)
                # Copy the formatting
                # openpyxl_sheet.cell(row=row_index+1, column=col_index+1).number_format = xlrd_workbook.xf_list[cell_format].format_str

    return openpyxl_workbook

def extract_info(wb, sheet_index):
    readout = {}
    sheet = wb[wb.sheetnames[sheet_index]]
    field_start = get_column_letter(sheet.min_column) + str(sheet.min_row)
    field_end = get_column_letter(sheet.max_column) + str(sheet.max_row)
    for rowOfCellObjects in sheet[field_start:field_end]:
        for cellObj in rowOfCellObjects:
            if cellObj.value is not None:
                if isinstance(cellObj.value, str) and cellObj.value.startswith("File No."):
                    readout["Filenumber"] = cellObj.value.replace("File No.", "").strip()
                elif cellObj.value == "Contractor":
                    coord_next = cellObj.coordinate.replace(get_column_letter(cellObj.col_idx), get_column_letter(cellObj.col_idx+2))
                    readout["Contractor"] = sheet[coord_next].value
                elif cellObj.value == "Contract No.":
                    coord_next = cellObj.coordinate.replace(get_column_letter(cellObj.col_idx), get_column_letter(cellObj.col_idx+2))
                    readout["Contract No"] = sheet[coord_next].value
                elif cellObj.value == "Project Name":
                    coord_next = cellObj.coordinate.replace(get_column_letter(cellObj.col_idx), get_column_letter(cellObj.col_idx+2))
                    readout["Project Name"] = sheet[coord_next].value
                elif cellObj.value == "Project No.":
                    coord_next = cellObj.coordinate.replace(get_column_letter(cellObj.col_idx), get_column_letter(cellObj.col_idx+1))
                    readout["Project No"] = sheet[coord_next].value
                elif cellObj.value == "Location":
                    coord_next = cellObj.coordinate.replace(get_column_letter(cellObj.col_idx), get_column_letter(cellObj.col_idx+1))
                    readout["Location"] = sheet[coord_next].value
                elif cellObj.value == "Department":
                    coord_next = cellObj.coordinate.replace(get_column_letter(cellObj.col_idx), get_column_letter(cellObj.col_idx+1))
                    readout["Department"] = sheet[coord_next].value
                elif cellObj.value == "S/Total":
                    coord_next = cellObj.coordinate.replace(get_column_letter(cellObj.col_idx), get_column_letter(cellObj.col_idx+1))
                    readout["Currency"] = sheet[coord_next].value
                elif isinstance(cellObj.value, str) and cellObj.value.startswith("Accumulated paid Amount"):
                    coord_next = cellObj.coordinate.replace(get_column_letter(cellObj.col_idx), get_column_letter(cellObj.col_idx+9))
                    coord_next = coord_next.replace(str(cellObj.row), str(cellObj.row+2))
                    readout["Accumulated paid Amount"] = round(float(sheet[coord_next].value), 2)
                elif isinstance(cellObj.value, str) and cellObj.value.startswith("Beneficiary company name"):
                    coord_next = cellObj.coordinate.replace(get_column_letter(cellObj.col_idx), get_column_letter(cellObj.col_idx+3))
                    readout["Beneficiary company name"] = sheet[coord_next].value
                elif isinstance(cellObj.value, str) and cellObj.value.startswith("Beneficiary company address"):
                    coord_next = cellObj.coordinate.replace(get_column_letter(cellObj.col_idx), get_column_letter(cellObj.col_idx+3))
                    readout["Beneficiary company address"] = sheet[coord_next].value
                elif isinstance(cellObj.value, str) and cellObj.value == "Beneficiary bank account":
                    coord_next = cellObj.coordinate.replace(get_column_letter(cellObj.col_idx), get_column_letter(cellObj.col_idx+3))
                    readout["Beneficiary bank account"] = sheet[coord_next].value
                elif isinstance(cellObj.value, str) and cellObj.value.startswith("Beneficiary bank account"):
                    coord_next = cellObj.coordinate.replace(get_column_letter(cellObj.col_idx), get_column_letter(cellObj.col_idx+3))
                    readout["Beneficiary bank account"] = sheet[coord_next].value
                elif isinstance(cellObj.value, str) and cellObj.value.startswith("SWIFT"):
                    coord_next = cellObj.coordinate.replace(get_column_letter(cellObj.col_idx), get_column_letter(cellObj.col_idx+3))
                    readout["SWIFT"] = sheet[coord_next].value                            
                elif isinstance(cellObj.value, str) and cellObj.value.startswith("Bank name"):
                    coord_next = cellObj.coordinate.replace(get_column_letter(cellObj.col_idx), get_column_letter(cellObj.col_idx+3))
                    readout["Bank name"] = sheet[coord_next].value

    return readout

if __name__ == '__main__':
    psr = argparse.ArgumentParser()
    psr.add_argument('-dir', type=str, required=True)
    args = psr.parse_args()

    if not exists(args.dir):
        exit(f"Directory \"{args.dir}\" does not exist")
    
    if not isdir(args.dir):
        exit(f"{args.dir} is not a directory") 

    column_widths = {"Filenumber": 10, "Project Name": 30, "Contractor": 30, "Currency": 10, "Beneficiary company address": 40, "Beneficiary bank account": 30}

    # every file that ends with '.xls*' under "sheets" dir
    files = glob.glob(args.dir + '/**/*.xls*', recursive=True)
    readouts = []
    for f in files:
        try:
            if f.endswith(".xlsx"):
                wb = load_workbook(f, data_only=True)
                readout = extract_info(wb, 0)
                pprint.pprint(readout)
                readouts.append(readout)

            elif f.endswith(".xls"):
                wb = xlrd.open_workbook(f, formatting_info=True)
                wb = convert_workbook_xlrd_to_openpyxl(wb)
                readout = extract_info(wb, 1)
                pprint.pprint(readout)
                readouts.append(readout)
            
        except Exception as e:
            print(f"Error: {e}")

    print(f"Exporting {len(readouts)} readouts...")

    wb = Workbook()
    ws = wb.active

    # Find the index of the list element with the largest length
    idx = max(range(len(readouts)), key=lambda i: len(readouts[i]))

    # fill-up the column names with key values from the first dictionary of the list
    for col, (key, value) in enumerate(readouts[idx].items(), start=1):
            column = get_column_letter(col)
            column_dimension = ws.column_dimensions[column]
            column_dimension.width = column_widths[key] if key in column_widths else 20
            ws["{}{}".format(column, 1)] = key

    for row, r in enumerate(readouts):
        for col, (key, value) in enumerate(r.items(), start=1):
            # find index of the column that has key value in the first row
            col_idx = None
            for column_index, cell in enumerate(ws[1]):
                if cell.value == key:
                    col_idx = column_index + 1
                    break
            if col_idx is not None:
                column = get_column_letter(col_idx)
                ws["{}{}".format(column, row+2)] = value
    
    wb.save("output.xlsx")
    print("done")
